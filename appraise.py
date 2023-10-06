import sys
import os.path
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font,PatternFill

OUTPUT_SHEET_NAME = "数据一览"
FILTER_SHEET_NAMES = [OUTPUT_SHEET_NAME, "Sheet1"]


def HYPERLINK(target, name):
    return '=HYPERLINK("#\'{}\'!A1", "{}")'.format(target, name)

def COUNTA(sheet_name, a, b, c, d, dif=1):
    return f"=(COUNTA('{sheet_name}'!{a}:{b}) - COUNTA('{sheet_name}'!{c}:{d}))/{dif}"

def COUNTA_S(sheet_name, a, b):
    return f"COUNTA('{sheet_name}'!{a}:{b})"

# 读取工作文件
def get_workbook(file_path:str) -> list | str:
    if not os.path.isfile(file_path):
        return "FILE_NOT_EXISIT"
    file_extension = os.path.splitext(file_path)[1]
    if file_extension.lower() in ['.xlsx']:
        try:
            return load_workbook(filename=file_path)
        except Exception as e:
            return "FILE_OPEN_ERROR"
    else:
        return "FILE_NOT_EXCEL"

# 查找对应的值
def find_value(ws, value) -> str:
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).find(value) > 0:
                return cell.coordinate

# 标准化SHEET
def simple_sheet_style(sheet, title_number) -> None:
    for i in range(1, sheet.max_row):
        if i == 1:
            sheet.row_dimensions[i].height = 30
        elif i <= title_number:
            sheet.row_dimensions[i].height = 20
        else:
            sheet.row_dimensions[i].height = 50
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 50
    sheet.sheet_view.zoomScale = 100

#设置文字居中
def set_text_center(*args:Alignment):
    for arg in args:
        arg.alignment = Alignment(horizontal='center', vertical='center')

def main() -> None:
    file_name = input("请输入评语表路径:")
    load_result = get_workbook(file_name)
    if isinstance(load_result, str):
        print("ERROR:" + load_result)
        return
    wb = load_result
    print("开始加载文件:" + file_name + "\n")

    # 创建数据部分
    wb.create_sheet(OUTPUT_SHEET_NAME, 0)
    ws = wb[OUTPUT_SHEET_NAME]
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 6
    ws.column_dimensions['E'].width = 6
    ws.column_dimensions['F'].width = 6
    ws.column_dimensions['G'].width = 6

    START_ROW_LOCATION_INDEX = 3
    row_index = START_ROW_LOCATION_INDEX
    for sheet in wb:
        sheet_title = sheet.title
        # 过滤不需要的sheet
        if sheet_title in FILTER_SHEET_NAMES:
            continue
        print("开始处理:" + sheet_title)

        #定义列位置
        name_row = ws.cell(row=row_index, column=1)
        tassk_row = ws.cell(row=row_index, column=2)
        finished_row = ws.cell(row=row_index, column=3)
        count_row = ws.cell(row=row_index, column=4)
        money_row = ws.cell(row=row_index, column=5)
        money_count_row = ws.cell(row=row_index, column=6)
        if_null_row = ws.cell(row=row_index, column=7)

        name_row.value = HYPERLINK(sheet_title, sheet_title)
        name_row.font = Font(size=15, color="000000")
        set_text_center(name_row, tassk_row, finished_row, count_row, money_row, money_count_row, if_null_row)

        # 根据不同的平台选择公式
        if sheet_title.find("问大家") >= 0:
            simple_sheet_style(sheet, 1)
            tassk_row.value = f'={COUNTA_S(sheet_title, "A3", "A1000")}'
            finished_row.value = f'={COUNTA_S(sheet_title, "C3", "C1000")}'
            fill_color = "A8EAE4"
        elif sheet_title.find("拼多多") >= 0:
            simple_sheet_style(sheet, 1)
            tassk_row.value = f'={COUNTA_S(sheet_title, "A2", "A1000")}'
            finished_row.value = f'={COUNTA_S(sheet_title, "C2", "C1000")}'
            fill_color = "F4B7BE"
        else:
            simple_sheet_style(sheet, 2)
            tassk_row.value = f'={COUNTA_S(sheet_title, "A3", "A1000")}'
            finished_row.value = f'={COUNTA_S(sheet_title, "D3", "D1000")}'
            fill_color = "F9CBAA"
        #填充颜色
        name_row.fill = PatternFill(fill_type="solid", fgColor=fill_color)

        count_row.value = f"={tassk_row.coordinate}-{finished_row.coordinate}"
        if_null_row.value = f'=IF({count_row.coordinate}<=0,"","N")'
        money_row.value = 2 #默认为2
        money_count_row.value = f"={finished_row.coordinate}*{money_row.coordinate}"

        # 找商品ID
        # id_coordinate = find_value(sheet, "编码")

        # 修复单元格
        sheet.merge_cells('A1:D1')
        sheet.unmerge_cells('A1:D1')
        # 标题
        sheet["D1"].value = sheet_title
        sheet["D1"].alignment = Alignment(horizontal='center', vertical='center')
        # 返回的链接
        sheet["C1"].value = HYPERLINK(OUTPUT_SHEET_NAME, "返回")
        sheet["C1"].font = Font(size=20, bold=True, color='FF0000')
        sheet["C1"].alignment = Alignment(horizontal='center', vertical='center')

        row_index = row_index + 1

    # 数据概况
    ws["A1"].value = "任务完成度"
    ws["B1"].value = "总数"
    ws["C1"].value = "已完成"
    ws["D1"].value = "未完成"
    ws["E1"].value = "单价"
    ws["F1"].value = "花销"
    ws["G1"].value = ""

    ws["A2"].value = F"=C2/B2"
    ws['A2'].number_format = "0.00%"
    ws["B2"].value = F"=SUM(B{START_ROW_LOCATION_INDEX}:B{row_index})"
    ws["C2"].value = F"=SUM(C{START_ROW_LOCATION_INDEX}:C{row_index})"
    ws["D2"].value = F"=SUM(D{START_ROW_LOCATION_INDEX}:D{row_index})"
    ws["E2"].value = 0
    ws["F2"].value = F"=C{row_index}*E{row_index}"

    set_text_center(ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['A2'], ws['B2'], ws['C2'], ws['D2'])

    new_file_name = file_name.replace(".xlsx", "_处理后") + ".xlsx"
    wb.save(new_file_name)
    print("\n保存文件:" + new_file_name + " 成功")


if __name__ == '__main__':
    main()
