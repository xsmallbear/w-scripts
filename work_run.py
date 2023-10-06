import sys
from openpyxl import Workbook, load_workbook
import os
import re
import subprocess

PATTERN = r'\d{4}(?=\D*$)'

class Colors:
    # 定义颜色编码
    RESET = "\033[0m"
    RED = "\033[31m"
    GREEN = "\033[32m"
    YELLOW = "\033[33m"
    BLUE = "\033[34m"


class Order:
    def __init__(self, operator, shop, time, order_number, order_id, amount, gift, expenditure_channel, note,
                 operation_phone, phone_number, product_name, keywords, jd_to_tb_id):
        self.operator = operator
        self.shop = shop
        self.time = time
        self.order_number = order_number
        self.order_id = order_id
        self.amount = amount
        self.gift = gift
        self.expenditure_channel = expenditure_channel
        self.note = note
        self.operation_phone = operation_phone
        self.phone_number = phone_number
        self.product_name = product_name
        self.keywords = keywords
        self.jd_to_tb_id = jd_to_tb_id

    def info(self):
        matches = re.findall(PATTERN, str(self.product_name))
        return f"操作人:\t\t\t{self.operator}\n" \
               f"店铺:\t\t\t{self.shop}\n" \
               f"时间:\t\t\t{self.time}\n" \
               f"客户订单编号:\t\t{self.order_number}\n" \
               f"客户ID:\t\t\t{self.order_id}\n" \
               f"备注:\t\t\t{self.note}\n" \
               f"操作手机:\t\t{self.operation_phone}\n" \
               f"产品名称:\t\t{self.product_name}\n" \
               f"关键词:\t\t\t{self.keywords}\n" \
               f"京东单对应淘宝ID:\t{self.jd_to_tb_id}\n" \
               f"产品名称:\t\t{Colors.RED}{matches}{Colors.RESET}"

#清屏命令
def clear_screen():
    os_type = os.name
    clear_command = ""
    if os_type == "posix":
        clear_command = "clear"
    elif os_type == "nt":
        clear_command = "cls"
    else:
        return
    os.system(clear_command)


def input_clipboard(text: str) -> None:
    subprocess.run(['echo', text, '|', 'clip'], shell=True)
    print(f"已将=>\n{text}\n=复制到剪切板\n")


# 读取工作文件
def get_workbook(file_path: str) -> Workbook | str:
    if not os.path.isfile(file_path):
        return "FILE_NOT_EXISIT"
    file_extension = os.path.splitext(file_path)[1]
    if file_extension.lower() in ['.xlsx']:
        try:
            return load_workbook(filename=file_path)
        except Exception as e:
            return "FILE_OPEN_ERROR"
    else:
        return "FILE_NOT_EXCEL"


# 从订单列列表中结果
def find_name_to_order_info(orderList: list[Order], current_id: str) -> list[Order]:
    results = []
    for order in orderList:
        if current_id == order.order_id:
            results.append(order)
    return results


def get_input() -> str:
    current_id = None
    while True:
        try:
            current_id = input("请输入(输入exit退出):")
            if current_id != "" and current_id is not None:
                return current_id.replace(" ", "")
        except KeyboardInterrupt:
            print()


def calc_phones(datas: list[Order]):
    phones = {}
    first = True
    for data in datas:
        if first is True:
            first = False
            continue
        current_phone = data.operation_phone
        if current_phone in phones:
            phones[current_phone] = phones[current_phone] + 1
        else:
            phones[current_phone] = 1
    phones = dict(sorted(phones.items(), key=lambda x: x[1], reverse=True))
    return phones


def print_phone_info(phones):
    print("=====手机信息=====")
    print(f"个数{len(phones.items())}\n")
    for string, count in phones.items():
        if string is None:
            string = "空白"
        print(f"{string}:{count}")
    print("\n=================\n")


def main():
    file_name = input("请输入工作表的地址:")
    load_result = get_workbook(file_name)
    if isinstance(load_result, str):
        print("ERROR:" + load_result)
        return
    wb = load_result
    ws = wb.worksheets[0]

    # 开局清屏
    clear_screen()
    orderList = []
    # 遍历单元格
    for row in range(1, ws.max_row + 1):
        order_values = []
        for col in range(1, 14 + 1):
            order_values.append(ws.cell(row=row, column=col).value)
        orderList.append(Order(*order_values))
    print(f"数据读取成完毕，一共读取了{len(orderList)}条数据\n")
    phones = calc_phones(orderList)
    print_phone_info(phones)

    while True:
        current_id = get_input()
        if current_id.upper() == "EXIT":
            print("感谢使用哦~~~~~~~~~")
            return
        if current_id.upper() == "PHONES":
            clear_screen()
            print_phone_info(phones)
            continue
        clear_screen()
        results = find_name_to_order_info(orderList, current_id)
        if 0 < len(results) < 2:
            result_str = ""
            result = results[0]
            print("------------------------------------------------")
            print(result.info())
            print("------------------------------------------------")
            result_str = result_str + f"{result.order_number} {result.order_id}"
            print(f"查询到数据结果{len(results)}条\n")
            input_clipboard(f"{result_str}")
            pass
        elif len(results) > 1:
            result_str = ""
            print("------------------------------------------------")
            print(results[0].info())
            print("------------------------------------------------")
            for result in results:
                result_str = result_str + f"{result.order_number} {result.order_id}\n"
            print(f"查询到数据结果{len(results)}条")
            # input_clipboard(f"{result_str}")
            print(result_str)
        else:
            print(f"\n{current_id}:没有数据~~~~~~~~~~~~~")
            continue

if __name__ == '__main__':
    main()
